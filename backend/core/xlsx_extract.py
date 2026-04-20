"""
XLSX data extractor using openpyxl.
No LibreOffice needed. Extracts directly from native format.
Each sheet becomes a "page" in page_content.
"""
from __future__ import annotations

from typing import Optional, Callable

from openpyxl import load_workbook

from backend.core import database as db


def extract_xlsx(file_path: str, sop_id: str, on_status: Optional[Callable] = None, tenant_id: str = None) -> dict:
    """
    Extract data from XLSX file.
    Each sheet = one "page" in page_content.
    """
    if on_status:
        on_status("xlsx", f"Extracting XLSX: {sop_id}")

    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        return {"error": f"Cannot open XLSX: {e}"}

    total_rows = 0
    sheet_count = 0

    for sheet_idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        sheet_count += 1
        page_num = sheet_idx + 1

        # Extract all rows as text
        text_lines = [f"**Sheet: {sheet_name}**\n"]
        headers = []
        rows_data = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            cells = [str(cell) if cell is not None else "" for cell in row]

            # Skip completely empty rows
            if not any(cells):
                continue

            if row_idx == 0:
                headers = cells
                text_lines.append("| " + " | ".join(cells) + " |")
                text_lines.append("|" + "|".join(["---"] * len(cells)) + "|")
            else:
                rows_data.append(cells)
                text_lines.append("| " + " | ".join(cells) + " |")
                total_rows += 1

        text = "\n".join(text_lines)

        # Build table JSON
        tables = []
        if headers:
            tables.append({"headers": headers, "rows": rows_data[:100]})  # Limit rows

        db.upsert_page_content(
            sop_id=sop_id,
            page=page_num,
            text_content=text,
            tables=tables,
            has_tables=True,
            extraction_method="xlsx",
            key_info=f"Sheet '{sheet_name}' with {len(rows_data)} data rows",
            tenant_id=tenant_id,
        )

    # Generate HTML preview
    try:
        preview_dir = db.DATA_DIR / "previews"
        preview_dir.mkdir(parents=True, exist_ok=True)
        html_parts = ["""<!DOCTYPE html><html><head><meta charset="utf-8">
        <style>body{font-family:Inter,sans-serif;padding:40px;max-width:1200px;margin:0 auto;color:#1a1a1a}
        table{border-collapse:collapse;width:100%;margin:16px 0}th,td{border:1px solid #e0e0e0;padding:8px 12px;text-align:left;font-size:13px}
        th{background:#f5f5f5;font-weight:700;position:sticky;top:0}tr:nth-child(even){background:#fafafa}
        h2{font-size:18px;margin:24px 0 8px;color:#333}</style></head><body>"""]
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            html_parts.append(f"<h2>{sheet_name}</h2><table>")
            for row_idx, row in enumerate(ws.iter_rows(values_only=True, max_row=500)):
                cells = [str(c) if c is not None else "" for c in row]
                if not any(cells):
                    continue
                tag = "th" if row_idx == 0 else "td"
                html_parts.append("<tr>" + "".join(f"<{tag}>{c}</{tag}>" for c in cells) + "</tr>")
            html_parts.append("</table>")
        html_parts.append("</body></html>")
        (preview_dir / f"{sop_id}.html").write_text("".join(html_parts), encoding="utf-8")
    except Exception:
        pass

    wb.close()

    if on_status:
        on_status("xlsx_done", f"XLSX done: {sheet_count} sheets, {total_rows} rows")

    return {
        "pages": sheet_count,
        "tables": sheet_count,
        "rows": total_rows,
    }
